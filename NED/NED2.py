"""
軌跡解析および可視化システム（GUI版）

複数のGPS軌跡データを読み込み、基準線を基準とした座標変換を行い、
軌跡グラフとcross_track_errorグラフを生成してExcelに出力する。
"""

import csv
import math
import sys
import os
from typing import List, Dict, Optional, Tuple
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pyproj
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import io
from PIL import Image
import random
import datetime
import matplotlib.font_manager as fm


# 設定定数
PROJECTION_LAT_CENTER = 35.0000
PROJECTION_LON_CENTER = 135.0000
REQUIRED_FIELDS = ['target_lat', 'target_lon', 'heading_deg', 'offset_deg']
FIGURE_SIZE = (12, 8)

# 日本語フォントの設定
def setup_japanese_font():
    """日本語フォントを設定"""
    font = 'Yu Gothic'
    plt.rcParams['font.family'] = font
    return font


class TrajectoryAnalyzer:
    """軌跡解析のメインクラス"""
    
    def __init__(self):
        """初期化"""
        self.proj = pyproj.Proj(proj='aeqd', lat_0=PROJECTION_LAT_CENTER, lon_0=PROJECTION_LON_CENTER)
        self.japanese_font = setup_japanese_font()
        
        # 日本語フォントが利用できない場合のラベル設定
        if self.japanese_font is None:
            self.labels = {
                'trajectory_title': 'Trajectory from Baseline',
                'x_label': 'Cross-track Distance from Baseline [m]',
                'y_label': 'Along-track Distance [m]',
                'error_title': 'Cross Track Error',
                'error_x_label': 'Time (Data Points)',
                'error_y_label': 'Cross Track Error [m]',
                'baseline': 'Baseline',
                'perpendicular': 'Perpendicular',
                'trajectory': 'Trajectory'
            }
        else:
            self.labels = {
                'trajectory_title': '基準線基準の軌跡',
                'x_label': '基準線からの横方向距離 [m]',
                'y_label': '基準線方向の距離 [m]',
                'error_title': 'Cross Track Error',
                'error_x_label': '時間 (データポイント)',
                'error_y_label': 'Cross Track Error [m]',
                'baseline': '基準線',
                'perpendicular': '垂直線',
                'trajectory': '軌跡'
            }
    
    def latlon_to_xy(self, lat: float, lon: float) -> Tuple[float, float]:
        """緯度経度をXY座標に変換"""
        return self.proj(lon, lat)
    
    def load_csv_data(self, file_path: str) -> List[Dict[str, str]]:
        """CSVファイルを読み込む"""
        try:
            with open(file_path, 'r', newline='', encoding='utf-8') as f:
                return list(csv.DictReader(f))
        except Exception as e:
            raise Exception(f"CSVファイル読み込みエラー: {e}")
    
    def filter_turn_status_data(self, data: List[Dict[str, str]]) -> List[Dict[str, str]]:
        """旋回状態のデータをフィルタリング"""
        # turn_statusが空でないデータがある場合のみフィルタリング
        has_turn_status = any(row.get('turn_status') for row in data)
        if not has_turn_status:
            return data
        
        turn_status_data = []
        found_0_to_3 = False
        
        for row in data:
            if not row.get('turn_status'):
                if found_0_to_3:  # すでに開始していれば追加
                    turn_status_data.append(row)
                continue
            
            try:
                status = int(row['turn_status'])
                if status == 0 and not found_0_to_3:
                    found_0_to_3 = True
                
                if found_0_to_3:
                    turn_status_data.append(row)
                    if status == 1:
                        # 直近5件にstatus=3があるかチェック
                        recent_statuses = [
                            int(prev_row['turn_status']) 
                            for prev_row in turn_status_data[-5:] 
                            if prev_row.get('turn_status') and prev_row['turn_status'].isdigit()
                        ]
                        if 3 in recent_statuses:
                            break
            except ValueError:
                if found_0_to_3:
                    turn_status_data.append(row)
                continue
        
        return turn_status_data if turn_status_data else data
    
    def extract_reference_data(self, data: List[Dict[str, str]]) -> Optional[Dict[str, float]]:
        """基準線算出に必要なデータを抽出"""
        for row in data:
            if all(row.get(field) and row[field].strip() for field in REQUIRED_FIELDS):
                try:
                    return {
                        'target_lat': float(row['target_lat']),
                        'target_lon': float(row['target_lon']),
                        'heading_deg': float(row['heading_deg']),
                        'offset_deg': float(row['offset_deg'])
                    }
                except ValueError:
                    continue
        return None
    
    def calculate_baseline_vectors(self, heading_deg: float, offset_deg: float) -> Tuple[float, float, float, float, float]:
        """基準線の方向ベクトルと垂直ベクトルを計算"""
        baseline_heading = heading_deg - offset_deg
        baseline_rad = math.radians(baseline_heading)
        
        baseline_unit_x = math.sin(baseline_rad)
        baseline_unit_y = math.cos(baseline_rad)
        baseline_perp_x = baseline_unit_y
        baseline_perp_y = -baseline_unit_x
        
        return baseline_unit_x, baseline_unit_y, baseline_perp_x, baseline_perp_y, baseline_heading
    
    def extract_trajectory_coordinates(self, data: List[Dict[str, str]]) -> List[Tuple[float, float]]:
        """軌跡データをXY座標に変換"""
        trajectory_xys = []
        for row in data:
            if row.get('lat') and row.get('lon'):
                try:
                    lat, lon = float(row['lat']), float(row['lon'])
                    x, y = self.latlon_to_xy(lat, lon)
                    trajectory_xys.append((x, y))
                except ValueError:
                    continue
        return trajectory_xys
    
    def transform_to_baseline_coordinate(self, trajectory_xys: List[Tuple[float, float]],
                                       target_x: float, target_y: float,
                                       baseline_unit_x: float, baseline_unit_y: float,
                                       baseline_perp_x: float, baseline_perp_y: float) -> List[Tuple[float, float]]:
        """基準線を基準とした座標系に変換"""
        transformed_trajectory = []
        
        for x, y in trajectory_xys:
            dx, dy = x - target_x, y - target_y
            transformed_x = dx * baseline_perp_x + dy * baseline_perp_y
            transformed_y = dx * baseline_unit_x + dy * baseline_unit_y
            transformed_trajectory.append((transformed_x, transformed_y))
        
        return transformed_trajectory
    
    def extract_cross_track_error_data(self, data: List[Dict[str, str]]) -> Tuple[List[float], List[float]]:
        """cross_track_errorデータとタイムスタンプを抽出"""
        timestamps = []
        errors = []
        
        for i, row in enumerate(data):
            if row.get('cross_track_error'):
                try:
                    error = float(row['cross_track_error'])
                    # タイムスタンプがない場合はインデックスを使用
                    timestamp = i if not row.get('timestamp') else i
                    timestamps.append(timestamp)
                    errors.append(error)
                except ValueError:
                    continue
        
        return timestamps, errors
    
    def analyze_file(self, file_path: str) -> Dict:
        """単一ファイルの解析を実行"""
        try:
            # データ読み込み
            data = self.load_csv_data(file_path)
            
            # 基準データの抽出
            reference_data = self.extract_reference_data(data)
            if reference_data is None:
                raise Exception("基準線算出に必要なデータが不足しています")
            
            # 基準線計算
            baseline_unit_x, baseline_unit_y, baseline_perp_x, baseline_perp_y, baseline_heading = \
                self.calculate_baseline_vectors(reference_data['heading_deg'], reference_data['offset_deg'])
            
            # ターゲット座標変換
            target_x, target_y = self.latlon_to_xy(reference_data['target_lat'], reference_data['target_lon'])
            
            # 軌跡データ処理
            turn_status_data = self.filter_turn_status_data(data)
            trajectory_xys = self.extract_trajectory_coordinates(turn_status_data)
            
            # 座標変換
            transformed_trajectory = self.transform_to_baseline_coordinate(
                trajectory_xys, target_x, target_y,
                baseline_unit_x, baseline_unit_y, baseline_perp_x, baseline_perp_y
            )
            
            # cross_track_errorデータ抽出
            timestamps, cross_track_errors = self.extract_cross_track_error_data(turn_status_data)
            
            return {
                'file_path': file_path,
                'transformed_trajectory': transformed_trajectory,
                'timestamps': timestamps,
                'cross_track_errors': cross_track_errors,
                'baseline_heading': baseline_heading,
                'success': True
            }
            
        except Exception as e:
            return {
                'file_path': file_path,
                'error': str(e),
                'success': False
            }


class TrajectoryGUI:
    """軌跡解析GUI"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("軌跡解析システム")
        self.root.geometry("1000x750")
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.analyzer = TrajectoryAnalyzer()
        self.file_paths = []
        self.analysis_results = []
        self.export_directory = ""
        self.graph_width = 6
        self.graph_height = 6
        
        self.setup_ui()
    
    def setup_ui(self):
        """UIを設定"""
        # メインフレーム
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 設定フレーム
        settings_frame = ttk.LabelFrame(main_frame, text="設定")
        settings_frame.pack(fill=tk.X, pady=(0, 10))
        
        # エクスポート先設定
        export_frame = ttk.Frame(settings_frame)
        export_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(export_frame, text="エクスポート先:").pack(side=tk.LEFT)
        self.export_label = ttk.Label(export_frame, text="未設定", foreground="red")
        self.export_label.pack(side=tk.LEFT, padx=(5, 10))
        ttk.Button(export_frame, text="フォルダを選択", command=self.select_export_directory).pack(side=tk.LEFT)
        
        # グラフサイズ設定
        size_frame = ttk.Frame(settings_frame)
        size_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(size_frame, text="グラフサイズ:").pack(side=tk.LEFT)
        ttk.Label(size_frame, text="幅:").pack(side=tk.LEFT, padx=(10, 5))
        self.width_var = tk.StringVar(value="6")
        width_spinbox = ttk.Spinbox(size_frame, from_=4, to=15, width=5, textvariable=self.width_var)
        width_spinbox.pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(size_frame, text="高さ:").pack(side=tk.LEFT, padx=(5, 5))
        self.height_var = tk.StringVar(value="6")
        height_spinbox = ttk.Spinbox(size_frame, from_=4, to=12, width=5, textvariable=self.height_var)
        height_spinbox.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(size_frame, text="サイズを適用", command=self.apply_graph_size).pack(side=tk.LEFT)
        
        # ファイル選択フレーム
        file_frame = ttk.LabelFrame(main_frame, text="CSVファイル選択")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(file_frame, text="CSVファイルを選択", command=self.select_files).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(file_frame, text="解析実行", command=self.run_analysis).pack(side=tk.LEFT, padx=5, pady=5)
        
        # ファイルリスト
        list_frame = ttk.LabelFrame(main_frame, text="選択されたファイル")
        list_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.file_listbox = tk.Listbox(list_frame, height=4)
        self.file_listbox.pack(fill=tk.X, padx=5, pady=5)
        
        # プログレスバー
        self.progress = ttk.Progressbar(main_frame, length=400, mode='determinate')
        self.progress.pack(pady=(0, 10))
        
        # 結果表示フレーム
        result_frame = ttk.LabelFrame(main_frame, text="解析結果")
        result_frame.pack(fill=tk.BOTH, expand=True)
        
        # ノートブック（タブ）
        self.notebook = ttk.Notebook(result_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # ステータスバー
        self.status_var = tk.StringVar()
        self.status_var.set("エクスポート先を設定してからファイルを選択してください")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.pack(fill=tk.X, pady=(10, 0))
    
    def on_closing(self):
        """ウィンドウが閉じられるときの処理"""
        try:
            # 開いているmatplotlibのfigureをすべて閉じる
            plt.close('all')
        except:
            pass
        
        # アプリケーションを確実に終了
        self.root.quit()
        self.root.destroy()
    
    def select_export_directory(self):
        """エクスポート先ディレクトリを選択"""
        directory = filedialog.askdirectory(title="エクスポート先フォルダを選択")
        if directory:
            self.export_directory = directory
            self.export_label.config(text=os.path.basename(directory), foreground="green")
            self.status_var.set(f"エクスポート先が設定されました: {directory}")
    
    def apply_graph_size(self):
        """グラフサイズを適用"""
        try:
            self.graph_width = int(self.width_var.get())
            self.graph_height = int(self.height_var.get())
            self.status_var.set(f"グラフサイズを {self.graph_width}x{self.graph_height} に設定しました")
        except ValueError:
            messagebox.showerror("エラー", "グラフサイズは数値で入力してください")
    
    def select_files(self):
        """CSVファイルを選択"""
        if not self.export_directory:
            messagebox.showwarning("警告", "先にエクスポート先を設定してください")
            return
        
        files = filedialog.askopenfilenames(
            title="CSVファイルを選択",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if files:
            self.file_paths = list(files)
            self.file_listbox.delete(0, tk.END)
            for file_path in self.file_paths:
                self.file_listbox.insert(tk.END, os.path.basename(file_path))
            self.status_var.set(f"{len(self.file_paths)}個のファイルが選択されました")
    
    def run_analysis(self):
        """解析を実行"""
        if not self.export_directory:
            messagebox.showwarning("警告", "エクスポート先が設定されていません")
            return
        
        if not self.file_paths:
            messagebox.showwarning("警告", "CSVファイルが選択されていません")
            return
        
        self.apply_graph_size()  # グラフサイズを更新
        
        self.analysis_results = []
        self.progress['maximum'] = len(self.file_paths)
        self.progress['value'] = 0
        
        # 既存のタブをクリア
        for tab in self.notebook.tabs():
            self.notebook.forget(tab)
        
        successful_results = []
        
        for i, file_path in enumerate(self.file_paths):
            self.status_var.set(f"解析中: {os.path.basename(file_path)}")
            self.root.update()
            
            result = self.analyzer.analyze_file(file_path)
            self.analysis_results.append(result)
            
            if result['success']:
                self.create_result_tab(result, i)
                successful_results.append(result)
            else:
                messagebox.showerror("エラー", f"ファイル '{os.path.basename(file_path)}' の解析エラー:\n{result['error']}")
            
            self.progress['value'] = i + 1
            self.root.update()
        
        # 自動的にExcelエクスポートを実行
        if successful_results:
            self.auto_export_to_excel(successful_results)
        
        self.status_var.set(f"解析完了: {len(successful_results)}個のファイルを処理しました")
    
    def create_result_tab(self, result, index):
        """結果タブを作成"""
        tab_frame = ttk.Frame(self.notebook)
        filename = os.path.basename(result['file_path'])
        self.notebook.add(tab_frame, text=f"{filename}")
        
        # グラフを作成
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
        
        # 軌跡グラフ
        if result['transformed_trajectory']:
            t_xs, t_ys = zip(*result['transformed_trajectory'])
            ax1.plot(t_xs, t_ys, '-', linewidth=2, color='blue', label=self.analyzer.labels['trajectory'])
            ax1.axvline(x=0, color='red', linestyle='--', linewidth=1, label=self.analyzer.labels['baseline'])
            ax1.axhline(y=0, color='green', linestyle=':', linewidth=1, label=self.analyzer.labels['perpendicular'])
        
        ax1.set_title(f"{self.analyzer.labels['trajectory_title']} - {filename}")
        ax1.set_xlabel(self.analyzer.labels['x_label'])
        ax1.set_ylabel(self.analyzer.labels['y_label'])
        ax1.axis('equal')
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        
        # Cross Track Errorグラフ
        if result['timestamps'] and result['cross_track_errors']:
            ax2.plot(result['timestamps'], result['cross_track_errors'], '-o', linewidth=2, markersize=3, color='red', label='Cross Track Error')
            ax2.axhline(y=0, color='black', linestyle='-', linewidth=0.5, alpha=0.7)
        
        ax2.set_title(f"{self.analyzer.labels['error_title']} - {filename}")
        ax2.set_xlabel(self.analyzer.labels['error_x_label'])
        ax2.set_ylabel(self.analyzer.labels['error_y_label'])
        ax2.legend()
        ax2.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        # Tkinterにグラフを埋め込み
        canvas = FigureCanvasTkAgg(fig, tab_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def auto_export_to_excel(self, successful_results):
        """解析結果を自動的にExcelにエクスポート"""
        try:
            # タイムスタンプ付きファイル名を生成
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"trajectory_analysis_{timestamp}.xlsx"
            excel_path = os.path.join(self.export_directory, excel_filename)
            
            wb = Workbook()
            
            # デフォルトシートを削除
            wb.remove(wb.active)
            
            # 軌跡用シート
            trajectory_sheet = wb.create_sheet("軌跡グラフ" if self.analyzer.japanese_font else "Trajectory")
            
            # Cross Track Error用シート
            error_sheet = wb.create_sheet("Cross Track Error")
            
            self.status_var.set("Excelファイルを生成中...")
            self.root.update()
            
            # 各結果のグラフを画像として保存してExcelに追加
            row_offset_traj = 1
            row_offset_error = 1
            
            for i, result in enumerate(successful_results):
                filename = os.path.basename(result['file_path'])
                
                # 軌跡グラフを生成
                fig1, ax1 = plt.subplots(1, 1, figsize=(self.graph_width, self.graph_height))
                if result['transformed_trajectory']:
                    t_xs, t_ys = zip(*result['transformed_trajectory'])
                    ax1.plot(t_xs, t_ys, '-', linewidth=2, color='blue', label=self.analyzer.labels['trajectory'])
                    ax1.axvline(x=0, color='red', linestyle='--', linewidth=1, label=self.analyzer.labels['baseline'])
                    ax1.axhline(y=0, color='green', linestyle=':', linewidth=1, label=self.analyzer.labels['perpendicular'])
                
                ax1.set_title(f"{self.analyzer.labels['trajectory_title']} - {filename}")
                ax1.set_xlabel(self.analyzer.labels['x_label'])
                ax1.set_ylabel(self.analyzer.labels['y_label'])
                ax1.axis('equal')
                ax1.legend()
                ax1.grid(True, alpha=0.3)
                plt.tight_layout()
                
                # 軌跡グラフを画像として保存
                img_buffer1 = io.BytesIO()
                fig1.savefig(img_buffer1, format='png', dpi=100, bbox_inches='tight')
                img_buffer1.seek(0)
                img1 = OpenpyxlImage(img_buffer1)
                
                # Excelに挿入（軌跡シート）
                cell_pos = f"{chr(ord('A') + (i % 3) * 10)}{row_offset_traj + (i // 3) * 35}"
                trajectory_sheet.add_image(img1, cell_pos)
                
                plt.close(fig1)
                
                # Cross Track Errorグラフを生成
                fig2, ax2 = plt.subplots(1, 1, figsize=(self.graph_width, self.graph_height))
                if result['timestamps'] and result['cross_track_errors']:
                    ax2.plot(result['timestamps'], result['cross_track_errors'], '-o', linewidth=2, markersize=3, color='red')
                    ax2.axhline(y=0, color='black', linestyle='-', linewidth=0.5, alpha=0.7)
                
                ax2.set_title(f"{self.analyzer.labels['error_title']} - {filename}")
                ax2.set_xlabel(self.analyzer.labels['error_x_label'])
                ax2.set_ylabel(self.analyzer.labels['error_y_label'])
                ax2.grid(True, alpha=0.3)
                plt.tight_layout()
                
                # Cross Track Errorグラフを画像として保存
                img_buffer2 = io.BytesIO()
                fig2.savefig(img_buffer2, format='png', dpi=100, bbox_inches='tight')
                img_buffer2.seek(0)
                img2 = OpenpyxlImage(img_buffer2)
                
                # Excelに挿入（エラーシート）
                cell_pos2 = f"{chr(ord('A') + (i % 3) * 10)}{row_offset_error + (i // 3) * 35}"
                error_sheet.add_image(img2, cell_pos2)
                
                plt.close(fig2)
            
            # Excelファイルを保存
            wb.save(excel_path)
            
            self.status_var.set(f"Excelファイルが保存されました: {excel_filename}")
            messagebox.showinfo("完了", f"Excelファイルが正常に保存されました:\n{excel_path}")
            
        except Exception as e:
            messagebox.showerror("エラー", f"Excelエクスポートエラー: {e}")


def main():
    """メイン関数"""
    try:
        root = tk.Tk()
        app = TrajectoryGUI(root)
        root.mainloop()
    except KeyboardInterrupt:
        pass
    finally:
        # プログラム終了時にmatplotlibのfigureをクリーンアップ
        plt.close('all')
        # Pythonプロセスを確実に終了
        import sys
        sys.exit(0)


if __name__ == "__main__":
    main()